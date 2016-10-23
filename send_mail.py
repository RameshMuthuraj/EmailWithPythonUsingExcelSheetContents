'''
Author : Bhishan Bhandari
bbhishan@gmail.com

The following script uses 
smtplib to connect to the gmail client.

uses email module to format email and make it clean.

openpyxl to read/write excell file. openpyxl is a third party module and needs to be installed via following command
pip install openpyxl

The excel file containing Firstname Email must be present in the directory the script is present in and the name of the file must be email_list.xlsx . If not make change in line 41.

The program assumes records are kept in excel sheet starting from second row, first row being the header. If not 
make change in line 87

The text file containing email content must be in the directory the script is present in and the name of the file must be email_content.txt. If not make change in line 44.

Replace subject in line 37

Replace email by your email (gmail) in line 38

Replace password for your gmail account in line 39


'''


import smtplib
import openpyxl
import sys
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText

subject = "Test SMTP"	#replace with your email subject here.
fromaddr = "bbhishan@gmail.com"	#your gmail account address.
pwd = "password"	#replace with password for your gmail account.
no_of_emails_at_each_run = 100 # how many recipients to send email to at each run.
wb = openpyxl.load_workbook("email_list.xlsx", read_only=False)	#loads the excel sheet containing FirstName, Email in read/write mode.
ws = wb.active	#selects the active sheet(basically the first sheet in a multi-sheet excel.)

with open("email_content.txt", "rb") as f:	#using python's standard method to open text file containing email body.
    content = f.read()	#read all the contents of the file as a single string.


def send_clean_email(server, toaddr, name, row_num): 
    '''
    Uses SMTP instance passed as parameter from main method to send email to the toaddr email address parameter
 passed in. Customizes email content by adding Hello name where name is the name of the email holder, passed in as
 a parameter. In case sending email failed, Failed message is updated in the excel sheet for the respective email.
    '''
    msg = MIMEMultipart()	#instantiates a Multipurpose Internet Mail Extension so as to support anything from ASCII to non ASCII characters and text as well as non text formats including audio, video, etc.
    msg['From'] = fromaddr	#setting sender's email address in the MIMEMultipart object.
    msg['To'] = toaddr		#setting recipient's email address in the MIMEMultipart object.
    msg['Subject'] = subject	#setting subject in the MIMEMultipart object.
 
    body = "Hello " + name + "\n\n" + content #Appending Grettings. In this case Hello followed by recipient's FirstName read from the excel file. Email content read from file is then appended to the end of grettings.
    msg.attach(MIMEText(body, 'plain'))	#Setting email content to send in the MIMEMultipart object.
    text = msg.as_string()	# Setting the text to be sent as a sting.
    try:
        server.sendmail(fromaddr, toaddr, text)	#Tries to send email to the recipient.
    except:
        ws.cell(row = row_num, column = 4).value = "Failed"	#In case sending email failed, writes Failed to the fourth column of that particular email row. 


def main():
    '''
    Tries to create a connection to the gmail smtp client. If failed exits the program. Tries to read logfile 
which contains value of the last row in the excel sheet to whom email has been sent in the previous run. If file 
not found, then it is the first run of the script. so, the value for the row_value is set to 2. Reads the next 100
 rows and for each First Name, email combination calls send_clean_email method. At the end of the script updates 
logfile with the value of last row read from excel sheet.
    '''
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)	#Gmail's smtp client address is smtp.gmail.com and can be accessed via port 587
        server.starttls()	#use TLS encryption standards before sending email, password for security reasons.
        server.login(fromaddr, pwd)	#login as a SMTP client
    except:
        sys.exit(1)	#In case connection was unable to be established, exit the program.
    try:
        with open('logfile.txt', 'rb') as f:	# if not the first run of the script, the last row number to which email was sent is stored in the logfile.txt from the previous run. 
            row_value = f.read()
            row_value = int(row_value)	
    except IOError:
        row_value = 1	#In case, file not found meaning it is the first run of the script, it sets the row value to 2 meaning the FirstName, email data starts from second row. Assuming first row contains headers.

    for i in range(row_value, row_value + no_of_emails_at_each_run):
        try:
            name = ws.cell(row = i, column = 1).value 	#get first name from excel for ith row
            email = ws.cell(row = i, column = 2).value	#get email from excel for ith row
            if name  is not None and email is not None: #check if any of the first name or email fields are blank
                send_clean_email(server, email, name, i) #calling send_clean_email to send email to the ith person in the excel sheet.
            else:
                break	#if value at the excel sheet is not present, means the data is finished.  
        except:
            print "could not get name and email data"    
    with open('logfile.txt', 'wb') as f:	#open logfile in write mode
        f.write(str(i))	#at the end of the program, write the last number of row in the excel sheet to whom email was sent onto logfile.txt

if __name__ == '__main__':
    main()	#entry point of the program.
